from typing import List
import openpyxl as xl
import openpyxl.worksheet.worksheet as worksheet
from .configs import get_data_by_keyname, ExcelDataCfg, ExcelSheetCfg, DataEnum
from .error import CustomException
from .logger import Logger


def _process_nickname():
    """
    익명A, 익명B, ... 순서대로 익명 닉네임을 생성하는 generator 함수
    """
    nickname_char = ord("A")
    while True:
        yield f"익명{chr(nickname_char)}"
        nickname_char += 1


GenNickname = _process_nickname()

ProcessFunc = {
    ExcelDataCfg.Nickname: lambda nickname: (
        next(GenNickname) if nickname is None else nickname
    ),
    ExcelDataCfg.OriginalLink: lambda link: None if link is None else link,
    ExcelDataCfg.KorLyricsLink: lambda link: None if link is None else link,
}
"""
엑셀 파일의 데이터를 가공하는 함수들을 저장하는 dict
"""


def _get_name_from_row(sheet: xl.Workbook.worksheets, row_idx: int) -> str:
    """
    엑셀 행 번호를 받아서 해당 행이 어떤 사람의 정보인지 반환하는 함수

    Args:
        sheet (xl.Workbook.worksheets): 엑셀 시트
        row_idx (int): 엑셀 행 번호

    Returns:
        str: 해당 행의 사람 이름
    """
    col_idx = 1
    value = sheet.cell(row=ExcelSheetCfg.KeyDataRow.value, column=col_idx).value
    while value is not None:
        # 데이터 시트의 부원 이름에 해당하는 부분은 무조건 적어둘 거라는 가정하에 작성
        if value == ExcelDataCfg.Name.keyname:
            return sheet.cell(row=row_idx, column=col_idx).value
        col_idx += 1
        value = sheet.cell(row=ExcelSheetCfg.KeyDataRow.value, column=col_idx).value
    raise CustomException(
        "해당 열에서 이름에 해당하는 항목을 찾을 수 없습니다. 이러면 안되는데..."
    )


def get_data_order(sheet: worksheet) -> List[DataEnum]:
    """
    엑셀에서 존재해야 하는 데이터들이 어떤 순서대로 정렬되어 있는지 반환하는 함수

    Args:
        sheet (worksheet): 엑셀 시트

    Returns:
        List[ExcelData]: 엑셀 시트의 순서대로 정렬된 데이터의 종류들
    """
    try:
        # 특정 key 값이 어떤 열에 존재하는지 확인하고, 해당 순서를 list로 만들기
        key_row_idx = ExcelSheetCfg.KeyDataRow.value
        ordered_data = []  # 정렬된 데이터를 저장할 배열

        col_idx = 1
        while sheet.cell(row=key_row_idx, column=col_idx).value is not None:
            # 열을 한 칸씩 띄우면서 데이터를 작성하는 미친 사람이 없다는 가정하에 작성
            keyname = sheet.cell(row=key_row_idx, column=col_idx).value
            data = get_data_by_keyname(ExcelDataCfg, keyname)
            if data is None:
                raise CustomException(
                    f"찾으려고 하는 {keyname}이 엑셀 시트에는 없습니다."
                )
            ordered_data.append(data)
            col_idx += 1

        return ordered_data
    except Exception as e:
        raise e


def get_data(sheet: worksheet, ordered_data: List[ExcelDataCfg]) -> List[dict]:
    """
    엑셀 파일에서 데이터를 가져오는 함수

    Args:
        sheet (worksheet): 엑셀 시트
        ordered_data (List[ExcelData]): 엑셀 시트의 순서대로 정렬된 데이터의 종류들

    Returns:
        List[dict]: 엑셀 시트에서 추출한 데이터
    """
    row_idx = ExcelSheetCfg.StartDataRow.value

    playlist_data = []

    # 이름이 더 이상 없을 때까지 열을 하나씩 내려가면서 데이터 추출
    while sheet.cell(row=row_idx, column=1).value is not None:
        data_dict = {}

        person_name = _get_name_from_row(sheet, row_idx)
        for idx, data in enumerate(ordered_data):
            value = sheet.cell(row=row_idx, column=idx + 1).value
            # 해당 데이터가 꼭 있어야 하는 데이터인지 확인
            if data.essential and value is None:
                # 만약 없다면 에러 발생
                raise CustomException(
                    f"{person_name}의 {data.ui_name}에 해당하는 정보가 없는 것 같습니다. 엑셀 파일을 확인해주세요."
                )
            # 데이터가 있다면 데이터의 종류에 따라서 데이터 가공
            if data in ProcessFunc:
                processed_value = ProcessFunc[data](value)
                data_dict[data.keyname] = processed_value
            else:
                data_dict[data.keyname] = value

        playlist_data.append(data_dict)
        Logger.info(f"{person_name}님의 정보를 저장했습니다.")
        row_idx += 1

    return playlist_data


if __name__ == "__main__":
    pass
